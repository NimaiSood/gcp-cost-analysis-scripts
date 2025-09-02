import os
import datetime
import multiprocessing as mp
from concurrent.futures import ProcessPoolExecutor, as_completed
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from google.cloud import billing_v1, monitoring_v3, compute_v1

# --- Configuration ---
# Your billing account ID
BILLING_ACCOUNT_ID = "01227B-3F83E7-AC2416"

# The threshold for "idle" CPU utilization (as a percentage, e.g., 5.0 for 5%)
IDLE_CPU_THRESHOLD_PERCENT = 5.0

# The duration to check for idle status (in minutes, 2 days = 2880 minutes)
IDLE_DURATION_MINUTES = 2880
        print(f"\n📦 Processing Batch {batch_num}/{len(project_batches)} ({len(project_batch)} projects)")
        
        batch_results = []
        
        # Process projects in this batch in parallel
        with ProcessPoolExecutor(max_workers=min(max_workers, len(project_batch))) as executor:idle" CPU utilization (as a percentage, e.g., 5.0 for 5%)
IDLE_CPU_THRESHOLD_PERCENT = 5.0

# The duration to check for idle status (in minutes, 2 days = 2880 minutes)
IDLE_DURATION_MINUTES = 2880

# --- Excel Report Generation ---

def generate_excel_report(all_results, billing_account_id, cpu_threshold, duration_minutes):
    """
    Generate a comprehensive Excel report with idle VM instances.
    """
    try:
        # Create timestamp for filename
        timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"idle_vms_report_{billing_account_id}_{timestamp}.xlsx"
        
        print(f"\n📊 Generating Excel report: {filename}")
        
        # Prepare data for different sheets
        idle_instances_data = []
        project_summary_data = []
        error_summary_data = []
        
        total_idle = 0
        total_instances = 0
        
        # Process results for each project
        for result in all_results:
            project_id = result['project_id']
            project_total = result['total_instances']
            project_idle = len(result['idle_instances'])
            project_errors = len(result['errors'])
            
            total_instances += project_total
            total_idle += project_idle
            
            # Add to project summary
            project_summary_data.append({
                'Project ID': project_id,
                'Total Instances': project_total,
                'Idle Instances': project_idle,
                'Idle Percentage': f"{(project_idle/project_total*100):.1f}%" if project_total > 0 else "0.0%",
                'Errors': project_errors,
                'Status': result['status']
            })
            
            # Add idle instances details
            for instance in result['idle_instances']:
                idle_instances_data.append({
                    'Project ID': project_id,
                    'Instance Name': instance['name'],
                    'Zone': instance['zone'],
                    'CPU Utilization (%)': f"{instance['cpu_utilization']*100:.2f}%",
                    'CPU Utilization (Raw)': instance['cpu_utilization'],
                    'Static IP Address': instance['static_ip'],
                    'Has Disks': instance['has_disks'],
                    'Analysis Date': datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                    'Analysis Period (Days)': duration_minutes / 60 / 24,
                    'CPU Threshold (%)': cpu_threshold,
                    'Potential Monthly Savings': "Calculate based on instance type"  # Placeholder
                })
            
            # Add errors
            for error in result['errors']:
                error_summary_data.append({
                    'Project ID': project_id,
                    'Error Description': error,
                    'Timestamp': datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                })
        
        # Create Excel workbook
        with pd.ExcelWriter(filename, engine='openpyxl') as writer:
            
            # Sheet 1: Executive Summary
            exec_summary_data = {
                'Metric': [
                    'Billing Account ID',
                    'Analysis Date',
                    'Analysis Period (Days)',
                    'CPU Threshold (%)',
                    'Total Projects Analyzed',
                    'Total VM Instances Scanned',
                    'Total Idle Instances Found',
                    'Overall Idle Rate (%)',
                    'Projects with Idle VMs',
                    'Total Errors Encountered'
                ],
                'Value': [
                    billing_account_id,
                    datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                    f"{duration_minutes / 60 / 24:.1f}",
                    f"{cpu_threshold}%",
                    len(all_results),
                    total_instances,
                    total_idle,
                    f"{(total_idle/total_instances*100):.1f}%" if total_instances > 0 else "0.0%",
                    len([r for r in all_results if len(r['idle_instances']) > 0]),
                    sum(len(r['errors']) for r in all_results)
                ]
            }
            exec_summary_df = pd.DataFrame(exec_summary_data)
            exec_summary_df.to_excel(writer, sheet_name='Executive Summary', index=False)
            
            # Sheet 2: Idle Instances Details
            if idle_instances_data:
                idle_instances_df = pd.DataFrame(idle_instances_data)
                idle_instances_df.to_excel(writer, sheet_name='Idle Instances', index=False)
            else:
                # Create empty sheet with headers
                empty_df = pd.DataFrame(columns=[
                    'Project ID', 'Instance Name', 'Zone', 'CPU Utilization (%)',
                    'Static IP Address', 'Has Disks', 'Analysis Date'
                ])
                empty_df.to_excel(writer, sheet_name='Idle Instances', index=False)
            
            # Sheet 3: Project Summary
            project_summary_df = pd.DataFrame(project_summary_data)
            project_summary_df.to_excel(writer, sheet_name='Project Summary', index=False)
            
            # Sheet 4: Error Summary
            if error_summary_data:
                error_summary_df = pd.DataFrame(error_summary_data)
                error_summary_df.to_excel(writer, sheet_name='Error Summary', index=False)
            else:
                # Create empty sheet with headers
                empty_error_df = pd.DataFrame(columns=['Project ID', 'Error Description', 'Timestamp'])
                empty_error_df.to_excel(writer, sheet_name='Error Summary', index=False)
        
        # Apply formatting to the Excel file
        format_excel_report(filename)
        
        print(f"✅ Excel report generated successfully: {filename}")
        print(f"📋 Report contains {len(idle_instances_data)} idle instances across {len(project_summary_data)} projects")
        return filename
        
    except Exception as e:
        print(f"❌ Error generating Excel report: {e}")
        import traceback
        traceback.print_exc()
        return None

def format_excel_report(filename):
    """
    Apply formatting to the Excel report for better readability.
    """
    try:
        workbook = openpyxl.load_workbook(filename)
        
        # Define styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        center_alignment = Alignment(horizontal='center', vertical='center')
        
        # Format each sheet
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            
            # Format headers (first row)
            for cell in sheet[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_alignment
                cell.border = border
            
            # Auto-adjust column widths
            for column in sheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                adjusted_width = min(max_length + 2, 50)  # Max width of 50
                sheet.column_dimensions[column_letter].width = adjusted_width
            
            # Add borders to all cells with data
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.value is not None:
                        cell.border = border
        
        # Special formatting for Idle Instances sheet
        if 'Idle Instances' in workbook.sheetnames:
            idle_sheet = workbook['Idle Instances']
            
            # Highlight high CPU utilization in red
            red_fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
            yellow_fill = PatternFill(start_color="FFFACD", end_color="FFFACD", fill_type="solid")
            
            # Find CPU utilization column
            cpu_col = None
            for col in range(1, idle_sheet.max_column + 1):
                if idle_sheet.cell(1, col).value == 'CPU Utilization (Raw)':
                    cpu_col = col
                    break
            
            if cpu_col:
                for row in range(2, idle_sheet.max_row + 1):
                    cell = idle_sheet.cell(row, cpu_col)
                    if cell.value is not None:
                        try:
                            cpu_value = float(cell.value)
                            if cpu_value > 0.03:  # 3%
                                cell.fill = yellow_fill
                            elif cpu_value > 0.01:  # 1%
                                cell.fill = red_fill
                        except:
                            pass
        
        workbook.save(filename)
        print(f"✅ Excel formatting applied successfully")
        
    except Exception as e:
        print(f"⚠️  Warning: Could not apply Excel formatting: {e}")

# --- Main Script Logic ---

def process_single_project(project_id, billing_account_id, cpu_threshold, duration_minutes):
    """
    Process a single project to find idle VMs.
    This function is designed to be run in parallel.
    """
    try:
        # Initialize clients for this process
        compute_client = compute_v1.InstancesClient()
        monitoring_client = monitoring_v3.MetricServiceClient()
        
        results = {
            'project_id': project_id,
            'idle_instances': [],
            'total_instances': 0,
            'errors': [],
            'status': 'processing'
        }
        
        print(f"[{project_id}] Starting project analysis...")
        
        # Get static IPs for this project
        static_ips = set()
        try:
            addresses_client = compute_v1.AddressesClient()
            aggregated_addresses = addresses_client.aggregated_list(project=project_id)
            for region, addresses_scoped_list in aggregated_addresses:
                if addresses_scoped_list.addresses:
                    for addr in addresses_scoped_list.addresses:
                        if addr.status == compute_v1.Address.Status.RESERVED:
                            static_ips.add(addr.address)
        except Exception as addr_error:
            results['errors'].append(f"Could not fetch addresses: {addr_error}")
            static_ips = set()

        # Get all instances in the project
        try:
            aggregated_list = compute_client.aggregated_list(project=project_id)
            
            for zone, scope in aggregated_list:
                if scope.instances:
                    results['total_instances'] += len(scope.instances)
                    zone_name = zone.split('/')[-1] if '/' in zone else zone
                    
                    for instance in scope.instances:
                        try:
                            # Check for disks
                            has_active_disk = bool(instance.disks)
                            
                            # Check for static IP
                            has_static_ip = False
                            static_ip_address = None
                            try:
                                for network_interface in instance.network_interfaces:
                                    for access_config in network_interface.access_configs:
                                        external_ip = access_config.nat_i_p
                                        if external_ip and external_ip in static_ips:
                                            has_static_ip = True
                                            static_ip_address = external_ip
                                            break
                                    if has_static_ip:
                                        break
                            except Exception as ip_error:
                                results['errors'].append(f"Could not check static IP for {instance.name}: {ip_error}")
                                has_static_ip = False

                            # Check CPU utilization
                            is_idle = False
                            avg_utilization = 0
                            
                            try:
                                end_time = datetime.datetime.now(tz=datetime.timezone.utc)
                                start_time = end_time - datetime.timedelta(minutes=duration_minutes)

                                from google.protobuf import timestamp_pb2
                                
                                end_timestamp = timestamp_pb2.Timestamp()
                                end_timestamp.FromDatetime(end_time)
                                
                                start_timestamp = timestamp_pb2.Timestamp()
                                start_timestamp.FromDatetime(start_time)
                                
                                request = monitoring_v3.ListTimeSeriesRequest(
                                    name=f"projects/{project_id}",
                                    filter=f'metric.type="compute.googleapis.com/instance/cpu/utilization" AND resource.labels.instance_id="{instance.id}"',
                                    interval=monitoring_v3.TimeInterval(
                                        end_time=end_timestamp,
                                        start_time=start_timestamp
                                    ),
                                    view=monitoring_v3.ListTimeSeriesRequest.TimeSeriesView.FULL
                                )

                                response = monitoring_client.list_time_series(request=request)
                                
                                total_utilization = 0
                                point_count = 0
                                
                                for time_series in response:
                                    for point in time_series.points:
                                        total_utilization += point.value.double_value
                                        point_count += 1
                                
                                if point_count > 0:
                                    avg_utilization = total_utilization / point_count
                                    if avg_utilization < (cpu_threshold / 100):
                                        is_idle = True
                                else:
                                    results['errors'].append(f"No monitoring data for {instance.name}")
                                    continue
                                    
                            except Exception as monitoring_error:
                                results['errors'].append(f"Monitoring error for {instance.name}: {monitoring_error}")
                                continue

                            # Check if instance meets all criteria
                            if is_idle and has_active_disk and has_static_ip:
                                idle_instance = {
                                    'name': instance.name,
                                    'zone': zone_name,
                                    'cpu_utilization': avg_utilization,
                                    'static_ip': static_ip_address,
                                    'has_disks': has_active_disk
                                }
                                results['idle_instances'].append(idle_instance)
                                print(f"[{project_id}] ✅ IDLE INSTANCE FOUND: {instance.name}")
                                print(f"    Zone: {zone_name}")
                                print(f"    CPU: {avg_utilization:.2%}")
                                print(f"    Static IP: {static_ip_address}")
                                print(f"    Has Disks: {has_active_disk}")
                                print("    " + "-" * 50)
                                
                        except Exception as instance_error:
                            results['errors'].append(f"Error processing instance {getattr(instance, 'name', 'unknown')}: {instance_error}")
                            continue
                            
        except Exception as compute_error:
            results['errors'].append(f"Failed to list instances: {compute_error}")
            results['status'] = 'failed'
            return results

        results['status'] = 'completed'
        if len(results['idle_instances']) > 0:
            print(f"[{project_id}] ✅ Completed: {len(results['idle_instances'])} idle instances found")
        return results
        
    except Exception as project_error:
        return {
            'project_id': project_id,
            'idle_instances': [],
            'total_instances': 0,
            'errors': [f"Project processing failed: {project_error}"],
            'status': 'failed'
        }

def list_idle_instances(billing_account_id, cpu_threshold, duration_minutes, batch_size=100, max_workers=20):
    """
    Lists compute instances that are idle for a specific duration, have active disks,
    and are assigned a static IP. Processes projects in parallel batches.
    """
    try:
        billing_client = billing_v1.CloudBillingClient()
        print("✅ Successfully initialized Google Cloud billing client")
    except Exception as e:
        print(f"❌ Failed to initialize Google Cloud billing client: {e}")
        import traceback
        traceback.print_exc()
        return

    billing_account_name = f"billingAccounts/{billing_account_id}"
    
    # 1. Get all projects associated with the billing account
    print(f"🔎 Scanning projects under billing account: {billing_account_id}...\n")
    try:
        print("Attempting to list project billing info...")
        # Create request to list projects for this billing account
        request = billing_v1.ListProjectBillingInfoRequest(
            name=billing_account_name
        )
        
        # Get all projects that are linked to this billing account
        response = billing_client.list_project_billing_info(request=request)
        all_projects = [
            project.project_id
            for project in response
            if project.billing_enabled
        ]
        
        # Process ALL projects (removed the [:10] limit)
        projects = all_projects
        print(f"Found {len(all_projects)} total projects linked to billing account")
        print(f"Processing ALL {len(projects)} projects in parallel batches of {batch_size}")
        print(f"🎯 ONLY IDLE INSTANCES WILL BE DISPLAYED")
    except Exception as e:
        print(f"❌ Error listing projects for billing account '{billing_account_id}': {e}")
        print(f"Exception type: {type(e).__name__}")
        import traceback
        traceback.print_exc()
        return

    # 2. Process projects in parallel batches
    all_results = []
    total_idle_instances = 0
    total_instances_scanned = 0
    total_errors = 0
    
    # Split projects into batches
    project_batches = [projects[i:i + batch_size] for i in range(0, len(projects), batch_size)]
    
    print(f"\n🚀 Starting parallel processing of {len(project_batches)} batches...")
    print(f"Using up to {max_workers} parallel workers per batch")
    print("=" * 80)
    
    for batch_num, project_batch in enumerate(project_batches, 1):
        print(f"\n� Processing Batch {batch_num}/{len(project_batches)} ({len(project_batch)} projects)")
        print(f"Projects in this batch: {', '.join(project_batch[:5])}{'...' if len(project_batch) > 5 else ''}")
        
        batch_results = []
        
        # Process projects in this batch in parallel
        with ProcessPoolExecutor(max_workers=min(max_workers, len(project_batch))) as executor:
            # Submit all projects in this batch
            future_to_project = {
                executor.submit(process_single_project, project_id, billing_account_id, cpu_threshold, duration_minutes): project_id
                for project_id in project_batch
            }
            
            # Collect results as they complete
            for future in as_completed(future_to_project):
                project_id = future_to_project[future]
                try:
                    result = future.result(timeout=300)  # 5-minute timeout per project
                    batch_results.append(result)
                    
                    # Update counters
                    total_idle_instances += len(result['idle_instances'])
                    total_instances_scanned += result['total_instances']
                    total_errors += len(result['errors'])
                    
                    # Only print if idle instances were found
                    if len(result['idle_instances']) > 0:
                        print(f"  🎯 {project_id}: {len(result['idle_instances'])} IDLE INSTANCES FOUND!")
                    
                except Exception as e:
                    print(f"  ❌ {project_id}: Failed with error: {e}")
                    batch_results.append({
                        'project_id': project_id,
                        'idle_instances': [],
                        'total_instances': 0,
                        'errors': [str(e)],
                        'status': 'failed'
                    })
                    total_errors += 1
        
        all_results.extend(batch_results)
        
        # Batch summary
        batch_idle = sum(len(r['idle_instances']) for r in batch_results)
        batch_total = sum(r['total_instances'] for r in batch_results)
        print(f"  📊 Batch {batch_num} Summary: {batch_idle} idle instances found out of {batch_total} total")
    
    # 3. Generate Excel Report
    excel_filename = generate_excel_report(all_results, billing_account_id, cpu_threshold, duration_minutes)
    
    # 4. Print final results
    print("\n" + "=" * 80)
    print("🏁 FINAL RESULTS")
    print("=" * 80)
    
    if total_idle_instances > 0:
        print(f"\n🎯 Found {total_idle_instances} idle instances across {len(projects)} projects:")
        print("-" * 80)
        
        for result in all_results:
            if result['idle_instances']:
                print(f"\n📍 Project: {result['project_id']}")
                for instance in result['idle_instances']:
                    print(f"  ✅ {instance['name']} (Zone: {instance['zone']})")
                    print(f"     - CPU Utilization: {instance['cpu_utilization']:.2%} (over {duration_minutes} mins)")
                    print(f"     - Static IP: {instance['static_ip']}")
                    print(f"     - Has Disks: {instance['has_disks']}")
    else:
        print(f"\n🎉 No idle instances found that meet all criteria across {len(projects)} projects!")
    
    # Summary statistics
    print(f"\n📈 SUMMARY STATISTICS:")
    print(f"  - Total Projects Processed: {len(projects)}")
    print(f"  - Total VM Instances Scanned: {total_instances_scanned}")
    print(f"  - Idle Instances Found: {total_idle_instances}")
    print(f"  - Total Errors Encountered: {total_errors}")
    print(f"  - Success Rate: {((len(projects) - total_errors) / len(projects) * 100):.1f}%")
    
    if excel_filename:
        print(f"\n📊 EXCEL REPORT GENERATED:")
        print(f"  - File: {excel_filename}")
        print(f"  - Contains: Detailed analysis across 4 worksheets")
        print(f"  - Executive Summary, Idle Instances, Project Summary, Error Summary")
    
    if total_errors > 0:
        print(f"\n⚠️  ERRORS SUMMARY:")
        for result in all_results:
            if result['errors']:
                print(f"  {result['project_id']}: {len(result['errors'])} errors")
                for error in result['errors'][:3]:  # Show first 3 errors
                    print(f"    - {error}")
                if len(result['errors']) > 3:
                    print(f"    - ... and {len(result['errors']) - 3} more errors")

if __name__ == "__main__":
    try:
        print("🚀 Starting Parallel Idle VM Scanner...")
        print(f"Configuration:")
        print(f"  - Billing Account: {BILLING_ACCOUNT_ID}")
        print(f"  - CPU Threshold: {IDLE_CPU_THRESHOLD_PERCENT}%")
        print(f"  - Duration: {IDLE_DURATION_MINUTES} minutes ({IDLE_DURATION_MINUTES/60/24:.1f} days)")
        print(f"  - Processing: ALL projects in parallel batches")
        print(f"  - Batch Size: 100 projects per batch")
        print(f"  - Max Workers: 20 parallel processes per batch")
        print(f"  - CPU Cores Available: {mp.cpu_count()}")
        print(f"  - Output Mode: IDLE INSTANCES ONLY")
        print("=" * 80)
        
        list_idle_instances(
            billing_account_id=BILLING_ACCOUNT_ID, 
            cpu_threshold=IDLE_CPU_THRESHOLD_PERCENT, 
            duration_minutes=IDLE_DURATION_MINUTES,
            batch_size=100,
            max_workers=20
        )
        
        print("\n" + "=" * 80)
        print("🏁 Parallel scan completed successfully!")
        
    except KeyboardInterrupt:
        print("\n⚠️  Scan interrupted by user (Ctrl+C)")
    except Exception as main_error:
        print(f"\n❌ Unexpected error in main execution: {main_error}")
        import traceback
        traceback.print_exc()