#!/usr/bin/env python3
"""
GCP Unused Resources Finder
Finds unattached persistent disks, unused IP addresses, outdated snapshots,
and unaccessed storage buckets across all projects under a billing account.
"""

import os
import datetime
import logging
import multiprocessing as mp
from concurrent.futures import ProcessPoolExecutor, as_completed
from typing import Dict, List, Any, Optional
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

# Google Cloud imports
from google.cloud import billing_v1, compute_v1, storage, monitoring_v3
from google.api_core import exceptions
from google.protobuf import timestamp_pb2
import time
import traceback

# --- Configuration ---
BILLING_ACCOUNT_ID = "01227B-3F83E7-AC2416"
SNAPSHOT_AGE_DAYS = 30  # Consider snapshots older than this as outdated
BUCKET_INACTIVE_DAYS = 90  # Consider buckets inactive if no access for this many days
MAX_WORKERS = 10  # Number of parallel workers for processing projects
BATCH_SIZE = 500  # Process projects in batches to manage memory and time
MAX_RETRIES = 3  # Maximum number of retries for API calls
RETRY_DELAY = 5  # Delay between retries in seconds
API_TIMEOUT = 60  # Timeout for API calls in seconds

# Set up logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

def retry_api_call(func, *args, **kwargs):
    """Retry API calls with exponential backoff."""
    for attempt in range(MAX_RETRIES):
        try:
            return func(*args, **kwargs)
        except exceptions.ResourceExhausted as e:
            if attempt < MAX_RETRIES - 1:
                wait_time = RETRY_DELAY * (2 ** attempt)
                logger.warning(f"Rate limit exceeded, waiting {wait_time}s before retry {attempt + 1}")
                time.sleep(wait_time)
                continue
            else:
                logger.error(f"Rate limit exceeded after {MAX_RETRIES} attempts: {e}")
                raise
        except exceptions.DeadlineExceeded as e:
            if attempt < MAX_RETRIES - 1:
                wait_time = RETRY_DELAY * (2 ** attempt)
                logger.warning(f"Timeout occurred, waiting {wait_time}s before retry {attempt + 1}")
                time.sleep(wait_time)
                continue
            else:
                logger.error(f"Timeout after {MAX_RETRIES} attempts: {e}")
                raise
        except exceptions.InternalServerError as e:
            if attempt < MAX_RETRIES - 1:
                wait_time = RETRY_DELAY * (2 ** attempt)
                logger.warning(f"Internal server error, waiting {wait_time}s before retry {attempt + 1}")
                time.sleep(wait_time)
                continue
            else:
                logger.error(f"Internal server error after {MAX_RETRIES} attempts: {e}")
                raise
        except exceptions.Forbidden as e:
            logger.error(f"Access forbidden - insufficient permissions: {e}")
            raise
        except exceptions.NotFound as e:
            logger.warning(f"Resource not found: {e}")
            return None
        except Exception as e:
            if attempt < MAX_RETRIES - 1:
                wait_time = RETRY_DELAY * (2 ** attempt)
                logger.warning(f"Unexpected error, waiting {wait_time}s before retry {attempt + 1}: {e}")
                time.sleep(wait_time)
                continue
            else:
                logger.error(f"Unexpected error after {MAX_RETRIES} attempts: {e}")
                raise

def get_projects_under_billing_account(billing_account_id: str) -> List[str]:
    try:
        client = billing_v1.CloudBillingClient()
        billing_account_name = f'billingAccounts/{billing_account_id}'
        
        logger.info(f"Fetching projects under billing account: {billing_account_id}")
        projects = list(retry_api_call(client.list_project_billing_info, name=billing_account_name))
        
        project_ids = []
        for project in projects:
            if project.billing_enabled:
                project_ids.append(project.project_id)
        
        logger.info(f"Found {len(project_ids)} active projects")
        return project_ids
    
    except exceptions.PermissionDenied as e:
        logger.error(f"Permission denied accessing billing account {billing_account_id}: {e}")
        return []
    except exceptions.Forbidden as e:
        logger.error(f"Access forbidden to billing account {billing_account_id}: {e}")
        return []
    except Exception as e:
        logger.error(f"Error fetching projects: {e}")
        logger.error(f"Traceback: {traceback.format_exc()}")
        return [] outdated snapshots,
and unaccessed storage buckets across all projects under a billing account.
"""

import os
import datetime
import logging
import multiprocessing as mp
from concurrent.futures import ProcessPoolExecutor, as_completed
from typing import Dict, List, Any, Optional
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

# Google Cloud imports
from google.cloud import billing_v1, compute_v1, storage, monitoring_v3
from google.api_core import exceptions
from google.protobuf import timestamp_pb2
import time
import traceback

# --- Configuration ---
BILLING_ACCOUNT_ID = "01227B-3F83E7-AC2416"
SNAPSHOT_AGE_DAYS = 30  # Consider snapshots older than this as outdated
BUCKET_INACTIVE_DAYS = 90  # Consider buckets inactive if no access for this many days
MAX_WORKERS = 10  # Number of parallel workers for processing projects
BATCH_SIZE = 500  # Process projects in batches to manage memory and time
MAX_RETRIES = 3  # Maximum number of retries for API calls
RETRY_DELAY = 5  # Delay between retries in seconds
API_TIMEOUT = 60  # Timeout for API calls in seconds

# Set up logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

def retry_api_call(func, *args, **kwargs):
    """Retry API calls with exponential backoff."""
    for attempt in range(MAX_RETRIES):
        try:
            return func(*args, **kwargs)
        except exceptions.ResourceExhausted as e:
            if attempt < MAX_RETRIES - 1:
                wait_time = RETRY_DELAY * (2 ** attempt)
                logger.warning(f"Rate limit exceeded, waiting {wait_time}s before retry {attempt + 1}")
                time.sleep(wait_time)
                continue
            else:
                logger.error(f"Rate limit exceeded after {MAX_RETRIES} attempts: {e}")
                raise
        except exceptions.DeadlineExceeded as e:
            if attempt < MAX_RETRIES - 1:
                wait_time = RETRY_DELAY * (2 ** attempt)
                logger.warning(f"Timeout occurred, waiting {wait_time}s before retry {attempt + 1}")
                time.sleep(wait_time)
                continue
            else:
                logger.error(f"Timeout after {MAX_RETRIES} attempts: {e}")
                raise
        except exceptions.InternalServerError as e:
            if attempt < MAX_RETRIES - 1:
                wait_time = RETRY_DELAY * (2 ** attempt)
                logger.warning(f"Internal server error, waiting {wait_time}s before retry {attempt + 1}")
                time.sleep(wait_time)
                continue
            else:
                logger.error(f"Internal server error after {MAX_RETRIES} attempts: {e}")
                raise
        except exceptions.Forbidden as e:
            logger.error(f"Access forbidden - insufficient permissions: {e}")
            raise
        except exceptions.NotFound as e:
            logger.warning(f"Resource not found: {e}")
            return None
        except Exception as e:
            if attempt < MAX_RETRIES - 1:
                wait_time = RETRY_DELAY * (2 ** attempt)
                logger.warning(f"Unexpected error, waiting {wait_time}s before retry {attempt + 1}: {e}")
                time.sleep(wait_time)
                continue
            else:
                logger.error(f"Unexpected error after {MAX_RETRIES} attempts: {e}")
                raise

def get_projects_under_billing_account(billing_account_id: str) -> List[str]:
    """Get all project IDs under the specified billing account."""
    try:
        client = billing_v1.CloudBillingClient()
        billing_account_name = f'billingAccounts/{billing_account_id}'
        
        logger.info(f"Fetching projects under billing account: {billing_account_id}")
        projects = list(client.list_project_billing_info(name=billing_account_name))
        
        project_ids = []
        for project in projects:
            if project.billing_enabled:
                project_ids.append(project.project_id)
        
        logger.info(f"Found {len(project_ids)} active projects")
        return project_ids
    
    except Exception as e:
        logger.error(f"Error fetching projects: {e}")
        return []

def find_unattached_disks(project_id: str) -> List[Dict[str, Any]]:
    """Find unattached persistent disks in a project."""
    unattached_disks = []
    
    try:
        disks_client = compute_v1.DisksClient()
        instances_client = compute_v1.InstancesClient()
        
        # Get all attached disk names
        attached_disk_names = set()
        try:
            aggregated_instances = retry_api_call(instances_client.aggregated_list, project=project_id)
            for zone, instances_scoped_list in aggregated_instances:
                if instances_scoped_list.instances:
                    for instance in instances_scoped_list.instances:
                        for disk in instance.disks:
                            if disk.source:
                                disk_name = disk.source.split('/')[-1]
                                attached_disk_names.add(disk_name)
        except exceptions.PermissionDenied:
            logger.warning(f"Permission denied accessing instances in {project_id}")
            return []
        except exceptions.Forbidden:
            logger.warning(f"Access forbidden to instances in {project_id}")
            return []
        except Exception as e:
            logger.warning(f"Error getting attached disks for {project_id}: {e}")
        
        # Check all disks
        try:
            aggregated_disks = retry_api_call(disks_client.aggregated_list, project=project_id)
            for zone, disks_scoped_list in aggregated_disks:
                if disks_scoped_list.disks:
                    zone_name = zone.split('/')[-1] if '/' in zone else zone
                    for disk in disks_scoped_list.disks:
                        if disk.name not in attached_disk_names:
                            unattached_disks.append({
                                'name': disk.name,
                                'zone': zone_name,
                                'size_gb': disk.size_gb,
                                'type': disk.type.split('/')[-1] if disk.type else 'unknown',
                                'creation_timestamp': disk.creation_timestamp,
                                'status': disk.status
                            })
        except exceptions.PermissionDenied:
            logger.warning(f"Permission denied accessing disks in {project_id}")
            return []
        except exceptions.Forbidden:
            logger.warning(f"Access forbidden to disks in {project_id}")
            return []
    
    except Exception as e:
        logger.error(f"Error finding unattached disks in {project_id}: {e}")
        logger.error(f"Traceback: {traceback.format_exc()}")
    
    return unattached_disks

def find_unused_static_ips(project_id: str) -> List[Dict[str, Any]]:
    """Find unused static IP addresses in a project."""
    unused_ips = []
    
    try:
        addresses_client = compute_v1.AddressesClient()
        
        # Get all static IP addresses
        try:
            aggregated_addresses = retry_api_call(addresses_client.aggregated_list, project=project_id)
            for region, addresses_scoped_list in aggregated_addresses:
                if addresses_scoped_list.addresses:
                    region_name = region.split('/')[-1] if '/' in region else region
                    for address in addresses_scoped_list.addresses:
                        # Check if IP is not in use (status is RESERVED but no users)
                        if (address.status == compute_v1.Address.Status.RESERVED and 
                            not address.users):
                            unused_ips.append({
                                'name': address.name,
                                'region': region_name,
                                'address': address.address,
                                'address_type': address.address_type,
                                'creation_timestamp': address.creation_timestamp,
                                'status': address.status
                            })
        except exceptions.PermissionDenied:
            logger.warning(f"Permission denied accessing addresses in {project_id}")
            return []
        except exceptions.Forbidden:
            logger.warning(f"Access forbidden to addresses in {project_id}")
            return []
    
    except Exception as e:
        logger.error(f"Error finding unused IPs in {project_id}: {e}")
        logger.error(f"Traceback: {traceback.format_exc()}")
    
    return unused_ips

def find_outdated_snapshots(project_id: str, max_age_days: int) -> List[Dict[str, Any]]:
    """Find snapshots older than specified days."""
    outdated_snapshots = []
    
    try:
        snapshots_client = compute_v1.SnapshotsClient()
        cutoff_date = datetime.datetime.now(datetime.timezone.utc) - datetime.timedelta(days=max_age_days)
        
        try:
            snapshots = retry_api_call(snapshots_client.list, project=project_id)
            for snapshot in snapshots:
                try:
                    # Parse creation timestamp
                    creation_time = datetime.datetime.fromisoformat(
                        snapshot.creation_timestamp.replace('Z', '+00:00')
                    )
                    
                    if creation_time < cutoff_date:
                        outdated_snapshots.append({
                            'name': snapshot.name,
                            'source_disk': snapshot.source_disk.split('/')[-1] if snapshot.source_disk else 'unknown',
                            'creation_timestamp': snapshot.creation_timestamp,
                            'disk_size_gb': snapshot.disk_size_gb,
                            'storage_bytes': snapshot.storage_bytes,
                            'status': snapshot.status,
                            'age_days': (datetime.datetime.now(datetime.timezone.utc) - creation_time).days
                        })
                except ValueError as e:
                    logger.warning(f"Error parsing snapshot {snapshot.name} timestamp: {e}")
                except Exception as e:
                    logger.warning(f"Error processing snapshot {snapshot.name}: {e}")
        except exceptions.PermissionDenied:
            logger.warning(f"Permission denied accessing snapshots in {project_id}")
            return []
        except exceptions.Forbidden:
            logger.warning(f"Access forbidden to snapshots in {project_id}")
            return []
    
    except Exception as e:
        logger.error(f"Error finding outdated snapshots in {project_id}: {e}")
        logger.error(f"Traceback: {traceback.format_exc()}")
    
    return outdated_snapshots

def find_unaccessed_storage_buckets(project_id: str, inactive_days: int) -> List[Dict[str, Any]]:
    """Find storage buckets that haven't been accessed recently."""
    unaccessed_buckets = []
    
    try:
        storage_client = storage.Client(project=project_id)
        cutoff_date = datetime.datetime.now(datetime.timezone.utc) - datetime.timedelta(days=inactive_days)
        
        try:
            buckets = retry_api_call(storage_client.list_buckets)
            for bucket in buckets:
                try:
                    # Get bucket metadata
                    bucket_info = retry_api_call(storage_client.get_bucket, bucket.name)
                    
                    # Check last access time using monitoring (if available)
                    # For now, we'll use creation time and updated time as proxies
                    last_activity = None
                    if hasattr(bucket_info, 'updated') and bucket_info.updated:
                        last_activity = bucket_info.updated
                    elif hasattr(bucket_info, 'time_created') and bucket_info.time_created:
                        last_activity = bucket_info.time_created
                    
                    if last_activity and last_activity < cutoff_date:
                        # Get bucket size (limited to avoid timeout)
                        total_objects = 0
                        total_size_bytes = 0
                        try:
                            blobs = bucket_info.list_blobs()
                            for blob in blobs:
                                total_objects += 1
                                if blob.size:
                                    total_size_bytes += blob.size
                                # Limit to avoid timeout on large buckets
                                if total_objects > 1000:
                                    break
                        except exceptions.Forbidden:
                            logger.warning(f"Access forbidden to bucket contents: {bucket.name}")
                        except Exception as e:
                            logger.warning(f"Error reading bucket contents {bucket.name}: {e}")
                        
                        unaccessed_buckets.append({
                            'name': bucket.name,
                            'location': bucket_info.location,
                            'storage_class': bucket_info.storage_class,
                            'creation_time': bucket_info.time_created.isoformat() if bucket_info.time_created else None,
                            'last_updated': last_activity.isoformat() if last_activity else None,
                            'total_objects': total_objects,
                            'total_size_bytes': total_size_bytes,
                            'days_since_activity': (datetime.datetime.now(datetime.timezone.utc) - last_activity).days if last_activity else None
                        })
                
                except exceptions.NotFound:
                    logger.warning(f"Bucket {bucket.name} not found (may have been deleted)")
                    continue
                except exceptions.Forbidden:
                    logger.warning(f"Access forbidden to bucket: {bucket.name}")
                    continue
                except Exception as e:
                    logger.warning(f"Error checking bucket {bucket.name}: {e}")
                    continue
        except exceptions.PermissionDenied:
            logger.warning(f"Permission denied accessing storage buckets in {project_id}")
            return []
        except exceptions.Forbidden:
            logger.warning(f"Access forbidden to storage buckets in {project_id}")
            return []
    
    except Exception as e:
        logger.error(f"Error finding unaccessed buckets in {project_id}: {e}")
        logger.error(f"Traceback: {traceback.format_exc()}")
    
    return unaccessed_buckets

def process_project(project_id: str) -> Dict[str, Any]:
    """Process a single project to find all unused resources."""
    logger.info(f"Processing project: {project_id}")
    
    result = {
        'project_id': project_id,
        'status': 'success',
        'error': None,
        'unattached_disks': [],
        'unused_ips': [],
        'outdated_snapshots': [],
        'unaccessed_buckets': []
    }
    
    try:
        # Find unattached disks
        try:
            result['unattached_disks'] = find_unattached_disks(project_id)
            logger.info(f"  Found {len(result['unattached_disks'])} unattached disks")
        except Exception as e:
            logger.error(f"  Error finding unattached disks: {e}")
            result['error'] = f"Disks error: {str(e)}"
        
        # Find unused static IPs
        try:
            result['unused_ips'] = find_unused_static_ips(project_id)
            logger.info(f"  Found {len(result['unused_ips'])} unused IPs")
        except Exception as e:
            logger.error(f"  Error finding unused IPs: {e}")
            if result['error']:
                result['error'] += f"; IPs error: {str(e)}"
            else:
                result['error'] = f"IPs error: {str(e)}"
        
        # Find outdated snapshots
        try:
            result['outdated_snapshots'] = find_outdated_snapshots(project_id, SNAPSHOT_AGE_DAYS)
            logger.info(f"  Found {len(result['outdated_snapshots'])} outdated snapshots")
        except Exception as e:
            logger.error(f"  Error finding outdated snapshots: {e}")
            if result['error']:
                result['error'] += f"; Snapshots error: {str(e)}"
            else:
                result['error'] = f"Snapshots error: {str(e)}"
        
        # Find unaccessed storage buckets
        try:
            result['unaccessed_buckets'] = find_unaccessed_storage_buckets(project_id, BUCKET_INACTIVE_DAYS)
            logger.info(f"  Found {len(result['unaccessed_buckets'])} unaccessed buckets")
        except Exception as e:
            logger.error(f"  Error finding unaccessed buckets: {e}")
            if result['error']:
                result['error'] += f"; Buckets error: {str(e)}"
            else:
                result['error'] = f"Buckets error: {str(e)}"
        
        # If we had any errors but still got some results, mark as partial success
        if result['error'] and (result['unattached_disks'] or result['unused_ips'] or 
                               result['outdated_snapshots'] or result['unaccessed_buckets']):
            result['status'] = 'partial'
        elif result['error']:
            result['status'] = 'error'
        
    except Exception as e:
        logger.error(f"Critical error processing project {project_id}: {e}")
        logger.error(f"Traceback: {traceback.format_exc()}")
        result['status'] = 'error'
        result['error'] = str(e)
    
    return result

def has_unused_resources(result: Dict[str, Any]) -> bool:
    """Check if a project has any unused resources."""
    return (len(result['unattached_disks']) > 0 or 
            len(result['unused_ips']) > 0 or 
            len(result['outdated_snapshots']) > 0 or 
            len(result['unaccessed_buckets']) > 0)

def process_project_batch(project_ids: List[str], batch_num: int, total_batches: int) -> List[Dict[str, Any]]:
    """Process a batch of projects and return only those with unused resources."""
    logger.info(f"Processing batch {batch_num}/{total_batches} ({len(project_ids)} projects)")
    
    batch_results = []
    projects_with_resources = []
    error_count = 0
    success_count = 0
    partial_count = 0
    
    with ProcessPoolExecutor(max_workers=MAX_WORKERS) as executor:
        future_to_project = {
            executor.submit(process_project, project_id): project_id 
            for project_id in project_ids
        }
        
        for future in as_completed(future_to_project):
            project_id = future_to_project[future]
            try:
                result = future.result(timeout=API_TIMEOUT * 2)  # Allow extra time for processing
                batch_results.append(result)
                
                # Count status types
                if result['status'] == 'success':
                    success_count += 1
                elif result['status'] == 'partial':
                    partial_count += 1
                else:
                    error_count += 1
                
                # Only keep projects with unused resources
                if has_unused_resources(result):
                    projects_with_resources.append(result)
                    status_icon = "✅" if result['status'] == 'success' else "⚠️" if result['status'] == 'partial' else "❌"
                    logger.info(f"{status_icon} {project_id}: Found unused resources - "
                              f"Disks: {len(result['unattached_disks'])}, "
                              f"IPs: {len(result['unused_ips'])}, "
                              f"Snapshots: {len(result['outdated_snapshots'])}, "
                              f"Buckets: {len(result['unaccessed_buckets'])}")
                else:
                    logger.info(f"⚪ {project_id}: No unused resources found")
                    
            except Exception as e:
                logger.error(f"Critical error processing project {project_id}: {e}")
                error_count += 1
                batch_results.append({
                    'project_id': project_id,
                    'status': 'error',
                    'error': f"Processing failed: {str(e)}",
                    'unattached_disks': [],
                    'unused_ips': [],
                    'outdated_snapshots': [],
                    'unaccessed_buckets': []
                })
    
    logger.info(f"Batch {batch_num} completed: {len(projects_with_resources)}/{len(project_ids)} projects have unused resources")
    logger.info(f"Batch {batch_num} status: {success_count} success, {partial_count} partial, {error_count} errors")
    return projects_with_resources

def format_size_bytes(size_bytes: int) -> str:
    """Format bytes into human readable format."""
    if size_bytes == 0:
        return "0 B"
    
    size_names = ["B", "KB", "MB", "GB", "TB"]
    i = 0
    while size_bytes >= 1024 and i < len(size_names) - 1:
        size_bytes /= 1024.0
        i += 1
    
    return f"{size_bytes:.2f} {size_names[i]}"

def generate_excel_report(all_results: List[Dict[str, Any]], billing_account_id: str) -> str:
    """Generate an Excel report with all unused resources."""
    timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"unused_resources_report_{billing_account_id}_{timestamp}.xlsx"
    
    logger.info(f"Generating Excel report: {filename}")
    
    # Prepare data for different sheets
    summary_data = []
    disks_data = []
    ips_data = []
    snapshots_data = []
    buckets_data = []
    
    total_unattached_disks = 0
    total_unused_ips = 0
    total_outdated_snapshots = 0
    total_unaccessed_buckets = 0
    
    for result in all_results:
        project_id = result['project_id']
        
        # Project summary - only include projects with unused resources
        if has_unused_resources(result):
            summary_data.append({
                'Project ID': project_id,
                'Status': result['status'],
                'Unattached Disks': len(result['unattached_disks']),
                'Unused IPs': len(result['unused_ips']),
                'Outdated Snapshots': len(result['outdated_snapshots']),
                'Unaccessed Buckets': len(result['unaccessed_buckets']),
                'Total Unused Resources': (len(result['unattached_disks']) + 
                                         len(result['unused_ips']) + 
                                         len(result['outdated_snapshots']) + 
                                         len(result['unaccessed_buckets'])),
                'Error': result.get('error', '')
            })
        
        # Unattached disks
        for disk in result['unattached_disks']:
            disks_data.append({
                'Project ID': project_id,
                'Disk Name': disk['name'],
                'Zone': disk['zone'],
                'Size (GB)': disk['size_gb'],
                'Type': disk['type'],
                'Created': disk['creation_timestamp'],
                'Status': disk['status']
            })
            total_unattached_disks += 1
        
        # Unused IPs
        for ip in result['unused_ips']:
            ips_data.append({
                'Project ID': project_id,
                'IP Name': ip['name'],
                'Region': ip['region'],
                'Address': ip['address'],
                'Type': ip['address_type'],
                'Created': ip['creation_timestamp'],
                'Status': ip['status']
            })
            total_unused_ips += 1
        
        # Outdated snapshots
        for snapshot in result['outdated_snapshots']:
            snapshots_data.append({
                'Project ID': project_id,
                'Snapshot Name': snapshot['name'],
                'Source Disk': snapshot['source_disk'],
                'Created': snapshot['creation_timestamp'],
                'Age (Days)': snapshot['age_days'],
                'Disk Size (GB)': snapshot['disk_size_gb'],
                'Storage Size': format_size_bytes(snapshot['storage_bytes']) if snapshot['storage_bytes'] else 'Unknown',
                'Status': snapshot['status']
            })
            total_outdated_snapshots += 1
        
        # Unaccessed buckets
        for bucket in result['unaccessed_buckets']:
            buckets_data.append({
                'Project ID': project_id,
                'Bucket Name': bucket['name'],
                'Location': bucket['location'],
                'Storage Class': bucket['storage_class'],
                'Created': bucket['creation_time'],
                'Last Updated': bucket['last_updated'],
                'Days Inactive': bucket['days_since_activity'],
                'Objects Count': bucket['total_objects'],
                'Total Size': format_size_bytes(bucket['total_size_bytes']) if bucket['total_size_bytes'] else 'Unknown'
            })
            total_unaccessed_buckets += 1
    
    # Create Excel workbook
    with pd.ExcelWriter(filename, engine='openpyxl') as writer:
        # Summary sheet
        if summary_data:
            summary_df = pd.DataFrame(summary_data)
            # Sort by total unused resources (descending)
            summary_df = summary_df.sort_values('Total Unused Resources', ascending=False)
            summary_df.to_excel(writer, sheet_name='Summary', index=False)
        
        # Individual resource sheets
        if disks_data:
            disks_df = pd.DataFrame(disks_data)
            disks_df.to_excel(writer, sheet_name='Unattached Disks', index=False)
        
        if ips_data:
            ips_df = pd.DataFrame(ips_data)
            ips_df.to_excel(writer, sheet_name='Unused IPs', index=False)
        
        if snapshots_data:
            snapshots_df = pd.DataFrame(snapshots_data)
            snapshots_df.to_excel(writer, sheet_name='Outdated Snapshots', index=False)
        
        if buckets_data:
            buckets_df = pd.DataFrame(buckets_data)
            buckets_df.to_excel(writer, sheet_name='Unaccessed Buckets', index=False)
    
    # Format the Excel file
    wb = openpyxl.load_workbook(filename)
    
    # Format each sheet
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        
        # Header formatting
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[column_letter].width = min(max_length + 2, 50)
    
    wb.save(filename)
    
    # Print summary
    print(f"\n📊 UNUSED RESOURCES REPORT")
    print(f"=" * 50)
    print(f"Billing Account: {billing_account_id}")
    print(f"Generated: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print(f"Report saved as: {filename}")
    print(f"\n📈 SUMMARY:")
    print(f"  • Projects with Unused Resources: {len(summary_data)}")
    print(f"  • Unattached Disks: {total_unattached_disks}")
    print(f"  • Unused Static IPs: {total_unused_ips}")
    print(f"  • Outdated Snapshots (>{SNAPSHOT_AGE_DAYS} days): {total_outdated_snapshots}")
    print(f"  • Unaccessed Buckets (>{BUCKET_INACTIVE_DAYS} days): {total_unaccessed_buckets}")
    
    return filename

def main():
    """Main function to orchestrate the unused resources scan."""
    print(f"🔍 GCP Unused Resources Finder")
    print(f"Scanning billing account: {BILLING_ACCOUNT_ID}")
    print(f"Configuration:")
    print(f"  • Snapshot age threshold: {SNAPSHOT_AGE_DAYS} days")
    print(f"  • Bucket inactivity threshold: {BUCKET_INACTIVE_DAYS} days")
    print(f"  • Batch size: {BATCH_SIZE} projects")
    print(f"  • Max workers per batch: {MAX_WORKERS}")
    
    # Get all projects under the billing account
    project_ids = get_projects_under_billing_account(BILLING_ACCOUNT_ID)
    
    if not project_ids:
        logger.error("No projects found or unable to access billing account")
        return
    
    print(f"\n🏗️  Found {len(project_ids)} total projects")
    
    # Calculate number of batches
    total_batches = (len(project_ids) + BATCH_SIZE - 1) // BATCH_SIZE
    print(f"📦 Processing in {total_batches} batches of {BATCH_SIZE} projects each")
    
    # Process projects in batches
    all_results_with_resources = []
    projects_processed = 0
    
    for batch_num in range(total_batches):
        start_idx = batch_num * BATCH_SIZE
        end_idx = min(start_idx + BATCH_SIZE, len(project_ids))
        batch_project_ids = project_ids[start_idx:end_idx]
        
        print(f"\n🔄 Starting batch {batch_num + 1}/{total_batches}")
        print(f"   Projects {start_idx + 1}-{end_idx} of {len(project_ids)}")
        
        batch_results = process_project_batch(batch_project_ids, batch_num + 1, total_batches)
        all_results_with_resources.extend(batch_results)
        projects_processed += len(batch_project_ids)
        
        # Print batch summary
        print(f"📊 Batch {batch_num + 1} Summary:")
        print(f"   • Projects processed: {len(batch_project_ids)}")
        print(f"   • Projects with unused resources: {len(batch_results)}")
        print(f"   • Total progress: {projects_processed}/{len(project_ids)} ({projects_processed/len(project_ids)*100:.1f}%)")
        
        if batch_results:
            print(f"   • Projects with resources in this batch:")
            for result in batch_results[:5]:  # Show first 5
                print(f"     - {result['project_id']}: "
                      f"D:{len(result['unattached_disks'])}, "
                      f"I:{len(result['unused_ips'])}, "
                      f"S:{len(result['outdated_snapshots'])}, "
                      f"B:{len(result['unaccessed_buckets'])}")
            if len(batch_results) > 5:
                print(f"     ... and {len(batch_results) - 5} more projects")
    
    # Generate summary report
    print(f"\n🎯 FINAL SUMMARY")
    print(f"=" * 60)
    print(f"Total projects scanned: {len(project_ids)}")
    print(f"Projects with unused resources: {len(all_results_with_resources)}")
    print(f"Projects without unused resources: {len(project_ids) - len(all_results_with_resources)}")
    print(f"Efficiency: {(len(project_ids) - len(all_results_with_resources))/len(project_ids)*100:.1f}% of projects are clean")
    
    if all_results_with_resources:
        # Count total resources
        total_disks = sum(len(r['unattached_disks']) for r in all_results_with_resources)
        total_ips = sum(len(r['unused_ips']) for r in all_results_with_resources)
        total_snapshots = sum(len(r['outdated_snapshots']) for r in all_results_with_resources)
        total_buckets = sum(len(r['unaccessed_buckets']) for r in all_results_with_resources)
        
        print(f"\n📈 UNUSED RESOURCES FOUND:")
        print(f"  • Unattached Disks: {total_disks}")
        print(f"  • Unused Static IPs: {total_ips}")
        print(f"  • Outdated Snapshots (>{SNAPSHOT_AGE_DAYS} days): {total_snapshots}")
        print(f"  • Unaccessed Buckets (>{BUCKET_INACTIVE_DAYS} days): {total_buckets}")
        
        print(f"\n📋 PROJECTS WITH UNUSED RESOURCES:")
        print("-" * 80)
        print(f"{'Project ID':<30} {'Disks':<6} {'IPs':<4} {'Snaps':<6} {'Buckets':<8} {'Status'}")
        print("-" * 80)
        
        # Sort by total unused resources (descending)
        sorted_results = sorted(all_results_with_resources, 
                               key=lambda x: (len(x['unattached_disks']) + 
                                            len(x['unused_ips']) + 
                                            len(x['outdated_snapshots']) + 
                                            len(x['unaccessed_buckets'])), 
                               reverse=True)
        
        for result in sorted_results:
            status_icon = "✅" if result['status'] == 'success' else "❌"
            print(f"{result['project_id']:<30} "
                  f"{len(result['unattached_disks']):<6} "
                  f"{len(result['unused_ips']):<4} "
                  f"{len(result['outdated_snapshots']):<6} "
                  f"{len(result['unaccessed_buckets']):<8} "
                  f"{status_icon}")
        
        # Generate Excel report only for projects with unused resources
        report_file = generate_excel_report(all_results_with_resources, BILLING_ACCOUNT_ID)
        print(f"\n✅ Report generated: {report_file}")
        print(f"📊 The report contains only the {len(all_results_with_resources)} projects with unused resources")
    else:
        print(f"\n🎉 Excellent! No unused resources found across all {len(project_ids)} projects!")
    
    print(f"\n✅ Scan completed!")

if __name__ == "__main__":
    main()